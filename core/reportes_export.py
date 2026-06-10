from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Equipo


def obtener_datos_reporte_equipos(anio):
    """Consolida gastos y correctivos por equipo para un año."""
    datos = []
    totales = {
        'correctivos': 0,
        'gasto_correctivos': Decimal('0'),
        'gasto_preventivos': Decimal('0'),
        'gasto_compras': Decimal('0'),
        'total_gastos': Decimal('0'),
    }

    for eq in Equipo.objects.filter(activo=True):
        correctivos_anio = eq.mantenimientos_correctivos.filter(fecha__year=anio)
        gc = correctivos_anio.aggregate(total=Sum('costo'))['total'] or Decimal('0')
        gp = (
            eq.mantenimientos_preventivos.filter(
                estado='ejecutado',
                fecha_ejecutada__year=anio,
            ).aggregate(total=Sum('costo'))['total'] or Decimal('0')
        )
        gco = eq.compras.filter(fecha_compra__year=anio).aggregate(total=Sum('costo'))['total'] or Decimal('0')
        num_correctivos = correctivos_anio.count()
        total = gc + gp + gco

        datos.append({
            'equipo': eq,
            'correctivos': num_correctivos,
            'gasto_correctivos': gc,
            'gasto_preventivos': gp,
            'gasto_compras': gco,
            'total_gastos': total,
        })

        totales['correctivos'] += num_correctivos
        totales['gasto_correctivos'] += gc
        totales['gasto_preventivos'] += gp
        totales['gasto_compras'] += gco
        totales['total_gastos'] += total

    datos.sort(key=lambda x: x['total_gastos'], reverse=True)
    return datos, totales


def _fecha_generacion():
    return timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')


def _moneda(valor):
    return f'S/ {valor:,.2f}'


def exportar_reporte_equipos_excel(anio, datos, totales):
    wb = Workbook()
    ws = wb.active
    ws.title = f'Reporte {anio}'

    azul = PatternFill('solid', fgColor='4E73DF')
    azul_claro = PatternFill('solid', fgColor='E8EEF9')
    amarillo = PatternFill('solid', fgColor='FFF3CD')
    gris = PatternFill('solid', fgColor='F8F9FC')
    borde = Border(
        left=Side(style='thin', color='D1D3E2'),
        right=Side(style='thin', color='D1D3E2'),
        top=Side(style='thin', color='D1D3E2'),
        bottom=Side(style='thin', color='D1D3E2'),
    )

    ws.merge_cells('A1:G1')
    ws['A1'] = 'REPORTE DE GASTOS Y CORRECTIVOS POR EQUIPO'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = azul
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Año: {anio}  |  Generado: {_fecha_generacion()}  |  Zona horaria: Perú (America/Lima)'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='5A5C69')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells('A3:G3')
    ws['A3'] = (
        'Resumen de mantenimientos correctivos, preventivos ejecutados y compras '
        'asociadas a cada equipo activo.'
    )
    ws['A3'].font = Font(name='Calibri', size=9, color='858796')
    ws['A3'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[3].height = 24

    encabezados = [
        'Equipo', 'Categoría', 'N° Correctivos',
        'Gasto correctivos (S/)', 'Gasto preventivos (S/)',
        'Gasto compras (S/)', 'Total gastos (S/)',
    ]
    fila_inicio = 5
    for col, texto in enumerate(encabezados, 1):
        celda = ws.cell(row=fila_inicio, column=col, value=texto)
        celda.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        celda.fill = azul
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border = borde

    for idx, d in enumerate(datos, fila_inicio + 1):
        fila = [
            d['equipo'].nombre,
            d['equipo'].get_categoria_display(),
            d['correctivos'],
            float(d['gasto_correctivos']),
            float(d['gasto_preventivos']),
            float(d['gasto_compras']),
            float(d['total_gastos']),
        ]
        alerta = d['correctivos'] >= 5 or d['total_gastos'] >= 5000
        for col, valor in enumerate(fila, 1):
            celda = ws.cell(row=idx, column=col, value=valor)
            celda.border = borde
            celda.font = Font(name='Calibri', size=10, bold=(col == 7))
            if alerta:
                celda.fill = amarillo
            elif idx % 2 == 0:
                celda.fill = gris
            if col == 3:
                celda.alignment = Alignment(horizontal='center')
            elif col >= 4:
                celda.number_format = '#,##0.00'
                celda.alignment = Alignment(horizontal='right')
            else:
                celda.alignment = Alignment(horizontal='left', vertical='center')

    fila_total = fila_inicio + len(datos) + 1
    totales_fila = [
        'TOTAL GENERAL', '',
        totales['correctivos'],
        float(totales['gasto_correctivos']),
        float(totales['gasto_preventivos']),
        float(totales['gasto_compras']),
        float(totales['total_gastos']),
    ]
    for col, valor in enumerate(totales_fila, 1):
        celda = ws.cell(row=fila_total, column=col, value=valor)
        celda.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='2E59D9')
        celda.border = borde
        if col >= 4:
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal='right')
        elif col == 3:
            celda.alignment = Alignment(horizontal='center')
        else:
            celda.alignment = Alignment(horizontal='left')

    anchos = [28, 22, 14, 18, 18, 18, 18]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + col)].width = ancho

    ws.freeze_panes = 'A6'
    ws.sheet_view.showGridLines = False

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_reporte_equipos_pdf(anio, datos, totales):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=f'Reporte equipos {anio}',
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        'TituloReporte',
        parent=estilos['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#4E73DF'),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    subtitulo = ParagraphStyle(
        'SubtituloReporte',
        parent=estilos['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#5A5C69'),
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    nota = ParagraphStyle(
        'NotaReporte',
        parent=estilos['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=8,
        textColor=colors.HexColor('#858796'),
        alignment=TA_CENTER,
        spaceAfter=14,
    )

    elementos = [
        Paragraph('REPORTE DE GASTOS Y CORRECTIVOS POR EQUIPO', titulo),
        Paragraph(
            f'Año fiscal: <b>{anio}</b> &nbsp;&nbsp;|&nbsp;&nbsp; '
            f'Generado: {_fecha_generacion()} &nbsp;&nbsp;|&nbsp;&nbsp; Perú (America/Lima)',
            subtitulo,
        ),
        Paragraph(
            'Equipos activos — correctivos, preventivos ejecutados y compras del período.',
            nota,
        ),
    ]

    encabezados = [
        'Equipo', 'Categoría', 'Correctivos',
        'Gasto correctivos', 'Gasto preventivos', 'Gasto compras', 'Total gastos',
    ]
    filas = [encabezados]
    for d in datos:
        filas.append([
            d['equipo'].nombre,
            d['equipo'].get_categoria_display(),
            str(d['correctivos']),
            _moneda(d['gasto_correctivos']),
            _moneda(d['gasto_preventivos']),
            _moneda(d['gasto_compras']),
            _moneda(d['total_gastos']),
        ])

    filas.append([
        'TOTAL GENERAL', '',
        str(totales['correctivos']),
        _moneda(totales['gasto_correctivos']),
        _moneda(totales['gasto_preventivos']),
        _moneda(totales['gasto_compras']),
        _moneda(totales['total_gastos']),
    ])

    tabla = Table(
        filas,
        colWidths=[6.2 * cm, 4.2 * cm, 2.2 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm],
        repeatRows=1,
    )

    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4E73DF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D3E2')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2E59D9')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ])

    for i, d in enumerate(datos, start=1):
        if d['correctivos'] >= 5 or d['total_gastos'] >= 5000:
            estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF3CD'))

    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)
    elementos.append(Spacer(1, 0.4 * cm))
    elementos.append(Paragraph(
        '<font size="7" color="#858796">'
        'Filas resaltadas: equipos con 5 o más correctivos o gasto total ≥ S/ 5,000. '
        'Sistema de Mantenimiento de Edificio.</font>',
        ParagraphStyle('Pie', alignment=TA_LEFT),
    ))

    doc.build(elementos)
    buffer.seek(0)
    return buffer
